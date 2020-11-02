import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook('transactions.xlsx')
# selecting the sheet
#sheet = wb['Sheet1']
# wb.create_sheet("Sheet2", 0)
# wb.save('transactions.xlsx')

# fetch value
sheet = wb['Sheet1']
# row values

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# selecting a value (5.95 *0.9)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'b6')


wb.save('trans.xlsx')
